import click
import openpyxl
import os
import django

os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'classproject.settings')
django.setup()

from onlineapp.models import *

from openpyxl import load_workbook


@click.group()
def grp():
    pass

@grp.command()
def load_college_data():
    wb = load_workbook(filename="students.xlsx")
    s = wb.get_sheet_by_name("Colleges")
    row_count = s.max_row
    col_count = s.max_column

    for i in range(2,row_count+1):
	    col_name = s.cell(row = i,column = 1).value
	    loc = s.cell(row=i,column=3).value
	    acrynm = s.cell(row = i,column = 2).value
	    contactt = s.cell(row = i,column = 4).value
	    c = College(name = col_name , location = loc , acronym = acrynm ,contact = contactt)
	    c.save()

@grp.command()
def load_student_data():
    wb = load_workbook(filename="students.xlsx")
    s = wb.get_sheet_by_name("Current")
    row_count = s.max_row
    col_count = s.max_column

    for i in range(2, row_count + 1):
        student_name = s.cell(row=i, column=1).value
        col_name = s.cell(row=i, column=2).value
        email_id = s.cell(row=i, column=3).value
        db = s.cell(row=i, column=4).value

        c = Student(name = student_name,dob = "2000-05-10",email = email_id,db_folder = db,dropped_out = False)
        clg_list = College.objects.all()

        for k in clg_list:
            if k.acronym == col_name:
                c.college = College.objects.get(pk = k.id)
                break
        c.save()
@grp.command()
def load_mock_data():
    wb = load_workbook(filename="marks.xlsx")
    s = wb.get_sheet_by_name("Sheet")
    row_count = s.max_row
    col_count = s.max_column

    for i in range(2, row_count + 1):
        student_details = s.cell(row=i, column=1).value
        prob1 = s.cell(row=i, column=2).value
        prob2 = s.cell(row=i, column=3).value
        prob3 = s.cell(row=i, column=4).value
        prob4 = s.cell(row=i, column=5).value
        totall = s.cell(row=i, column=6).value

        temp_list = student_details.split('_')
        student_name = temp_list[2]

        c = MockTest1()
        c.problem1 = prob1
        c.problem2 = prob2
        c.problem3 = prob3
        c.problem4 = prob4
        c.total = totall

        stu_list = Student.objects.all()
        flag = 0
        for k in stu_list:
            flag = 0
            try:
                if k.db_folder.lower() == student_name.lower():
                    flag = 1
                    c.student = Student.objects.get(pk = k.id)
                    c.save()
                    break
            except('IntegrityError','Error'):
                c.delete()


@grp.command()
def load_drop_data():
    wb = load_workbook(filename="students.xlsx")
    s = wb.get_sheet_by_name("Deletions")
    row_count = s.max_row
    col_count = s.max_column

    for i in range(2, row_count + 1):
        student_name = s.cell(row=i, column=1).value
        col_name = s.cell(row=i, column=2).value
        email_id = s.cell(row=i, column=3).value
        db = s.cell(row=i, column=4).value

        c = Student(name=student_name, dob="2000-05-10", email=email_id, db_folder=db, dropped_out=True)
        clg_list = College.objects.all()

        for k in clg_list:
            if k.acronym == col_name:
                c.college = College.objects.get(pk=k.id)
                break
        c.save()

def delete_all():
    c = MockTest1.objects.all()
    for x in c:
        x.delete()

if __name__ == "__main__":
    delete_all()